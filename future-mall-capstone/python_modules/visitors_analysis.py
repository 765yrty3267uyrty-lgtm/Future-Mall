#!/usr/bin/env python3
"""
Future Mall - Visitors Analysis Program
Analyzes mall visitor data read from an Excel (.xlsx) sheet, determines the
busiest / quietest day, generates a bar chart, and exports the results back
into Excel / CSV / JSON.
"""

import sys
import os
import csv
import json
from datetime import datetime
from typing import List, Dict, Tuple, Optional
from statistics import median

try:
    import matplotlib
    matplotlib.use("Agg")  # headless-safe so charts work anywhere
    import matplotlib.pyplot as plt
    HAS_MATPLOTLIB = True
except Exception:
    HAS_MATPLOTLIB = False

try:
    import openpyxl
    HAS_OPENPYXL = True
except Exception:
    HAS_OPENPYXL = False

# Add shared constants to path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'shared'))
from constants import VISITORS, BRAND, format_number, print_header, print_section

EXCEL_INPUT = os.path.join(os.path.dirname(__file__), "data", "visitors_data.xlsx")
CHART_OUTPUT = os.path.join(os.path.dirname(__file__), "data", "visitors_chart.png")
EXCEL_OUTPUT = os.path.join(os.path.dirname(__file__), "data", "visitors_analysis.xlsx")


class VisitorsAnalyzer:
    """Analyzes visitor data and generates statistics."""

    def __init__(self):
        self.days = VISITORS["days_of_week"]
        self.visitor_data: List[int] = []
        self.analysis_results: Dict = {}

    def validate_visitor_count(self, value: str) -> Optional[int]:
        """Validate visitor count input."""
        try:
            count = int(value.strip())
            if count < 0:
                print("Visitor count cannot be negative.")
                return None
            return count
        except ValueError:
            print("Invalid input. Please enter a whole number.")
            return None

    def collect_data(self) -> bool:
        """Collect visitor data for each day."""
        print_header("VISITOR DATA COLLECTION")
        print(f"\nEnter visitor counts for each day ({len(self.days)} days).")
        print("Enter 'q' at any time to quit.\n")

        self.visitor_data = []
        for day in self.days:
            while True:
                user_input = input(f"{day}: ").strip()
                if user_input.lower() in ('q', 'quit', 'exit'):
                    return False

                count = self.validate_visitor_count(user_input)
                if count is not None:
                    self.visitor_data.append(count)
                    break

        return True

    def calculate_statistics(self) -> Dict:
        """Calculate all statistics from visitor data."""
        if not self.visitor_data:
            return {}

        total = sum(self.visitor_data)
        avg = total / len(self.visitor_data)
        max_visitors = max(self.visitor_data)
        min_visitors = min(self.visitor_data)
        max_day = self.days[self.visitor_data.index(max_visitors)]
        min_day = self.days[self.visitor_data.index(min_visitors)]
        days_recorded = len(self.visitor_data)

        # Optional statistics
        med = median(self.visitor_data)
        visitor_range = max_visitors - min_visitors

        # Percentage change (max vs min)
        if min_visitors > 0:
            pct_change = ((max_visitors - min_visitors) / min_visitors) * 100
        else:
            pct_change = float('inf') if max_visitors > 0 else 0

        # Daily changes
        daily_changes = []
        for i in range(1, len(self.visitor_data)):
            if self.visitor_data[i - 1] > 0:
                change = ((self.visitor_data[i] - self.visitor_data[i - 1]) /
                          self.visitor_data[i - 1]) * 100
            else:
                change = float('inf') if self.visitor_data[i] > 0 else 0
            daily_changes.append({
                'from_day': self.days[i - 1],
                'to_day': self.days[i],
                'change': change
            })

        results = {
            'total': total,
            'average': round(avg, 2),
            'maximum': max_visitors,
            'minimum': min_visitors,
            'max_day': max_day,
            'min_day': min_day,
            'days_recorded': days_recorded,
            'median': med,
            'range': visitor_range,
            'pct_change_max_min': round(pct_change, 2),
            'daily_changes': daily_changes,
            'raw_data': self.visitor_data.copy(),
        }

        self.analysis_results = results
        return results

    def display_results_table(self) -> None:
        """Display results in a formatted table."""
        if not self.analysis_results:
            print("No data to display.")
            return

        r = self.analysis_results
        width = 50

        print_header("VISITOR STATISTICS REPORT")
        print(f"\n{'Metric':<30} {'Value':>18}")
        print("=" * width)
        print(f"{'Total Visitors':<30} {format_number(r['total']):>18}")
        print(f"{'Average Daily':<30} {format_number(r['average']):>18}")
        print(f"{'Maximum':<30} {format_number(r['maximum']):>12} ({r['max_day']})")
        print(f"{'Minimum':<30} {format_number(r['minimum']):>12} ({r['min_day']})")
        print(f"{'Days Recorded':<30} {format_number(r['days_recorded']):>18}")
        print("-" * width)
        print(f"{'Median':<30} {format_number(r['median']):>18}")
        print(f"{'Range (Max - Min)':<30} {format_number(r['range']):>18}")
        print(f"{'% Change (Max vs Min)':<30} {r['pct_change_max_min']:>17.2f}%")
        print("=" * width)

    def display_daily_breakdown(self) -> None:
        """Display day-by-day breakdown."""
        print_section("Daily Breakdown")
        print(f"{'Day':<12} {'Visitors':>10} {'Change':>12}")
        print("-" * 36)

        for i, (day, count) in enumerate(zip(self.days, self.visitor_data)):
            if i == 0:
                change_str = "—"
            else:
                prev = self.visitor_data[i - 1]
                if prev > 0:
                    pct = ((count - prev) / prev) * 100
                    change_str = f"{pct:+.1f}%"
                else:
                    change_str = "N/A"

            print(f"{day:<12} {format_number(count):>10} {change_str:>12}")

    def display_detailed_analysis(self) -> None:
        """Display comprehensive analysis."""
        if not self.analysis_results:
            return

        self.display_results_table()
        print()
        self.display_daily_breakdown()

        # Additional insights
        print_section("Insights")
        r = self.analysis_results

        if r['pct_change_max_min'] > 50:
            print(f"• High variability: {r['pct_change_max_min']:.1f}% difference between "
                  f"busiest ({r['max_day']}) and quietest ({r['min_day']}) day.")
        else:
            print(f"• Moderate variability: {r['pct_change_max_min']:.1f}% difference "
                  f"between peak and low days.")

        # Trend analysis
        increasing_days = sum(1 for c in r['daily_changes'] if c['change'] > 0)
        decreasing_days = sum(1 for c in r['daily_changes'] if c['change'] < 0)

        if increasing_days > decreasing_days:
            print("• Overall trend: Increasing visitor traffic through the week.")
        elif decreasing_days > increasing_days:
            print("• Overall trend: Decreasing visitor traffic through the week.")
        else:
            print("• Overall trend: Stable visitor traffic through the week.")

        # Peak analysis
        print(f"• Peak day: {r['max_day']} ({format_number(r['maximum'])} visitors)")
        print(f"• Lowest day: {r['min_day']} ({format_number(r['minimum'])} visitors)")

    def save_to_csv(self, filename: Optional[str] = None) -> str:
        """Save analysis results to CSV file."""
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"visitors_analysis_{timestamp}.csv"

        with open(filename, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)

            # Header
            writer.writerow(["Future Mall - Visitor Analysis Report"])
            writer.writerow(["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
            writer.writerow([])

            # Raw data
            writer.writerow(["Day", "Visitors"])
            for day, count in zip(self.days, self.visitor_data):
                writer.writerow([day, count])
            writer.writerow([])

            # Statistics
            r = self.analysis_results
            writer.writerow(["Metric", "Value"])
            writer.writerow(["Total Visitors", r['total']])
            writer.writerow(["Average Daily", r['average']])
            writer.writerow(["Maximum", f"{r['maximum']} ({r['max_day']})"])
            writer.writerow(["Minimum", f"{r['minimum']} ({r['min_day']})"])
            writer.writerow(["Days Recorded", r['days_recorded']])
            writer.writerow(["Median", r['median']])
            writer.writerow(["Range", r['range']])
            writer.writerow(["% Change (Max vs Min)", f"{r['pct_change_max_min']:.2f}%"])
            writer.writerow([])

            # Daily changes
            writer.writerow(["Day", "Visitors", "Change from Previous (%)"])
            for i, (day, count) in enumerate(zip(self.days, self.visitor_data)):
                if i == 0:
                    writer.writerow([day, count, "N/A"])
                else:
                    prev = self.visitor_data[i - 1]
                    if prev > 0:
                        pct = ((count - prev) / prev) * 100
                        writer.writerow([day, count, f"{pct:+.2f}"])
                    else:
                        writer.writerow([day, count, "N/A"])

        return filename

    def save_to_json(self, filename: Optional[str] = None) -> str:
        """Save analysis results to JSON file."""
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"visitors_analysis_{timestamp}.json"

        data = {
            "generated": datetime.now().isoformat(),
            "days": self.days,
            "visitor_counts": self.visitor_data,
            "statistics": self.analysis_results,
        }

        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)

        return filename

    # ------------------------------------------------------------------
    # Excel-based analysis (within the specification)
    # ------------------------------------------------------------------

    def load_from_excel(self, filename: Optional[str] = None) -> bool:
        """Read visitor counts from an Excel (.xlsx) sheet.

        Expects a sheet with two columns: "Day" and "Visitors".
        If no file is given, a real sample workbook (data/visitors_data.xlsx)
        containing a full mall week is loaded.
        """
        path = filename or EXCEL_INPUT
        if not os.path.exists(path):
            print(f"Excel file not found: {path}")
            return False
        if not HAS_OPENPYXL:
            print("openpyxl is required to read Excel files. pip install openpyxl")
            return False

        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb.active
            sheet_days: List[str] = []
            counts: List[int] = []
            for row in ws.iter_rows(min_row=1, values_only=True):
                if row is None or len(row) < 2:
                    continue
                day, count = row[0], row[1]
                if day is None or count is None:
                    continue
                label = str(day).strip()
                if label.lower() in ("day", "days"):
                    continue
                sheet_days.append(label)
                counts.append(int(count))

            if not counts:
                print("No visitor data rows found in the Excel sheet.")
                return False

            self.days = sheet_days
            self.visitor_data = counts
            print_header("DATA LOADED FROM EXCEL")
            print(f"File: {path}")
            print(f"Loaded {len(counts)} day(s) of mall visitor data.")
            return True
        except Exception as e:
            print(f"Failed to read Excel file: {e}")
            return False

    def save_to_excel(self, filename: Optional[str] = None) -> str:
        """Export the analysis results into an Excel workbook."""
        if filename is None:
            filename = EXCEL_OUTPUT
        if not HAS_OPENPYXL:
            print("openpyxl is required. pip install openpyxl")
            return ""

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Analysis"
        ws.append(["Future Mall - Visitor Analysis Report"])
        ws.append(["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        ws.append([])

        r = self.analysis_results
        ws.append(["Metric", "Value"])
        ws.append(["Total Visitors", r['total']])
        ws.append(["Average Daily", r['average']])
        ws.append(["Busiest Day", f"{r['max_day']} ({r['maximum']} visitors)"])
        ws.append(["Quietest Day", f"{r['min_day']} ({r['minimum']} visitors)"])
        ws.append(["Days Recorded", r['days_recorded']])
        ws.append(["Median", r['median']])
        ws.append(["Range (Max - Min)", r['range']])
        ws.append(["% Change (Max vs Min)", f"{r['pct_change_max_min']:.2f}%"])
        ws.append([])

        ws.append(["Day", "Visitors"])
        for day, count in zip(self.days, self.visitor_data):
            ws.append([day, count])

        os.makedirs(os.path.dirname(filename), exist_ok=True)
        wb.save(filename)
        return filename

    def generate_chart(self, filename: Optional[str] = None) -> str:
        """Generate a bar chart of the mall's daily visitor data."""
        if filename is None:
            filename = CHART_OUTPUT
        if not HAS_MATPLOTLIB:
            print("matplotlib is required to generate the chart.")
            return ""

        r = self.analysis_results
        days = self.days
        counts = self.visitor_data
        colors = [BRAND["colors"]["primary"]] * len(days)
        busy_idx = self.visitor_data.index(r['maximum'])
        quiet_idx = self.visitor_data.index(r['minimum'])
        colors[busy_idx] = BRAND["colors"]["success"]
        colors[quiet_idx] = BRAND["colors"]["danger"]

        fig, ax = plt.subplots(figsize=(10, 5))
        bars = ax.bar(days, counts, color=colors, edgecolor="white")
        ax.axhline(r['average'], color=BRAND["colors"]["accent"],
                   linestyle="--", linewidth=1.4,
                   label=f"Average: {r['average']:,.0f}")
        for bar, c in zip(bars, counts):
            ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 8,
                    f"{c:,}", ha="center", fontsize=9)
        ax.set_title("Future Mall - Weekly Visitor Traffic")
        ax.set_xlabel("Day")
        ax.set_ylabel("Visitors")
        ax.set_ylim(0, max(counts) * 1.15)
        ax.legend()
        ax.grid(axis="y", alpha=0.35)
        fig.tight_layout()
        os.makedirs(os.path.dirname(filename), exist_ok=True)
        fig.savefig(filename, dpi=150)
        plt.close(fig)
        return filename

    def analyze_from_excel(self, excel_file: Optional[str] = None) -> bool:
        """Full pipeline: load Excel -> calculate stats -> chart + report."""
        if not self.load_from_excel(excel_file):
            return False
        self.calculate_statistics()
        self.display_detailed_analysis()

        print_section("BUSIEST & QUIETEST DAY")
        r = self.analysis_results
        print(f"• Busiest day : {r['max_day']} with {format_number(r['maximum'])} visitors")
        print(f"• Quietest day: {r['min_day']} with {format_number(r['minimum'])} visitors")

        chart = self.generate_chart()
        print(f"\nChart saved to: {chart}")
        excel = self.save_to_excel()
        print(f"Excel saved to: {excel}")
        return True

    def run(self) -> None:
        """Main application loop."""
        print_header("FUTURE MALL - VISITORS ANALYSIS")
        print(f"\n{BRAND['name']} - {BRAND['slogan']}")

        while True:
            # Collect data
            if not self.collect_data():
                print("\nData collection cancelled.")
                break

            if len(self.visitor_data) < VISITORS["min_days"]:
                print(f"\nNeed at least {VISITORS['min_days']} day(s) of data.")
                continue

            # Analyze
            self.calculate_statistics()

            # Display results
            self.display_detailed_analysis()

            # Save options
            print_section("Save Options")
            print("1. Save as CSV")
            print("2. Save as JSON")
            print("3. Save both")
            print("4. Don't save")

            while True:
                choice = input("Choose option (1-4): ").strip()
                if choice in ('1', '2', '3', '4'):
                    break
                print("Please enter 1, 2, 3, or 4.")

            if choice == '1':
                filename = self.save_to_csv()
                print(f"\nSaved to: {filename}")
            elif choice == '2':
                filename = self.save_to_json()
                print(f"\nSaved to: {filename}")
            elif choice == '3':
                csv_file = self.save_to_csv()
                json_file = self.save_to_json()
                print(f"\nSaved to: {csv_file} and {json_file}")

            # Continue?
            print()
            again = input("Analyze another week? (y/N): ").strip().lower()
            if again != 'y':
                break

            print("\n" + "=" * 60 + "\n")


def main():
    """Entry point for the visitors analysis program.

    Defaults to analysing the mall visitor data stored in Excel
    (data/visitors_data.xlsx), produces a chart and an Excel report.
    """
    import argparse

    parser = argparse.ArgumentParser(description="Future Mall visitors analysis")
    parser.add_argument("--excel", help="Path to an Excel (.xlsx) visitor data file")
    parser.add_argument("--excel-out", help="Output path for the Excel report")
    args = parser.parse_args()

    analyzer = VisitorsAnalyzer()
    if args.excel_out:
        globals()["EXCEL_OUTPUT"] = args.excel_out
    if not analyzer.analyze_from_excel(getattr(args, "excel", None)):
        print("Falling back to manual data entry...")
        analyzer.run()


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\nProgram interrupted. Goodbye!")
        sys.exit(0)
    except Exception as e:
        print(f"\nAn error occurred: {e}")
        sys.exit(1)