#!/usr/bin/env python3
"""Generate the Future Mall sample visitor-data Excel sheet."""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                   "data", "visitors_data.xlsx")
os.makedirs(os.path.dirname(OUT), exist_ok=True)

# A realistic mall week: Friday/Saturday weekend rush, Monday slow.
DATA = [
    ("Day", "Visitors"),
    ("Monday", 630),
    ("Tuesday", 715),
    ("Wednesday", 690),
    ("Thursday", 920),
    ("Friday", 1450),
    ("Saturday", 1580),
    ("Sunday", 1240),
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Mall Visitors"

header_fill = PatternFill("solid", fgColor="2563EB")
header_font = Font(bold=True, color="FFFFFF")

for i, (day, visited) in enumerate(DATA, start=1):
    ws.cell(row=i, column=1, value=day)
    ws.cell(row=i, column=2, value=visited)
    if i == 1:
        for c in (1, 2):
            cell = ws.cell(row=i, column=c)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

for col, width in (("A", 16), ("B", 12)):
    ws.column_dimensions[col].width = width

wb.save(OUT)
print("Wrote", OUT)